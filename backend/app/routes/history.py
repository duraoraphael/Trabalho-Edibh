import json
import csv
from io import BytesIO, StringIO
from datetime import datetime, timezone
from datetime import date
from urllib.parse import quote
from typing import List
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from app.schemas import ReportResponse
from app.services.report_service import ReportService
from app.auth.dependencies import get_current_user
from app.services.pdf_service import PDFService

router = APIRouter()
service = ReportService()


def _load_evidence_entries(raw_value: str | list | None) -> list[dict]:
    if raw_value in (None, ""):
        return []
    parsed = raw_value
    if isinstance(raw_value, str):
        try:
            parsed = json.loads(raw_value) if raw_value else []
        except json.JSONDecodeError:
            return []
    if not isinstance(parsed, list):
        return []

    normalized: list[dict] = []
    for item in parsed:
        if isinstance(item, str):
            name = item.rsplit("/", 1)[-1] if "/" in item else item
            normalized.append(
                {
                    "id": name,
                    "name": name,
                    "web_url": item,
                    "size_bytes": None,
                    "mime_type": "application/octet-stream",
                }
            )
            continue
        if isinstance(item, dict):
            web_url = item.get("web_url") or item.get("webUrl") or ""
            name = item.get("name") or (web_url.rsplit("/", 1)[-1] if web_url else "arquivo")
            normalized.append(
                {
                    "id": item.get("id") or name,
                    "name": name,
                    "web_url": web_url,
                    "size_bytes": item.get("size_bytes"),
                    "mime_type": item.get("mime_type") or item.get("mimeType"),
                }
            )
    return normalized


def serialize_report(item: dict) -> dict:
    fields = item.get("fields", {})
    evidencias = _load_evidence_entries(fields.get("Evidencias", "[]"))

    return {
        "id": item.get("id"),
        "instalacao": fields.get("Instalacao", ""),
        "sistema": fields.get("Sistema", ""),
        "equipamento": fields.get("Equipamento", ""),
        "data": fields.get("Data", ""),
        "gerencia": fields.get("Gerencia", ""),
        "situacao_identificada": fields.get("SituacaoIdentificada", ""),
        "status": fields.get("Status", "Em análise"),
        "custom_fields": json.loads(fields.get("CamposCustomizados", "{}") or "{}") if isinstance(fields.get("CamposCustomizados", "{}"), str) else (fields.get("CamposCustomizados") or {}),
        "usuario_criacao": fields.get("UsuarioCriacao", ""),
        "data_criacao": fields.get("DataCriacao", ""),
        "usuario_alteracao": fields.get("UsuarioAlteracao", ""),
        "data_alteracao": fields.get("DataAlteracao", ""),
        "motivo_edicao": fields.get("MotivoEdicao", ""),
        "historico_alteracoes": json.loads(fields.get("HistoricoAlteracoes", "[]") or "[]"),
        "evidencias": evidencias,
    }


def _get_filtered_history(
    gerencia: str | None,
    start_date: date | None,
    end_date: date | None,
    search: str | None,
    sort_by: str,
    sort_dir: str,
) -> list[dict]:
    filters: dict[str, str] = {}
    if gerencia:
        filters["gerencia"] = gerencia
    if start_date:
        filters["start_date"] = start_date.isoformat()
    if end_date:
        filters["end_date"] = end_date.isoformat()
    if search:
        filters["search"] = search

    raw = service.list_reports(filters)
    items = [serialize_report(item) for item in raw.get("value", [])]

    allowed_sort_fields = {"data", "instalacao", "gerencia", "equipamento", "usuario_criacao", "data_criacao"}
    normalized_sort_by = sort_by if sort_by in allowed_sort_fields else "data"
    reverse = sort_dir != "asc"
    return sorted(items, key=lambda item: str(item.get(normalized_sort_by, "")).lower(), reverse=reverse)


@router.get("/", response_model=List[ReportResponse])
def list_history(
    gerencia: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    search: str | None = None,
    sort_by: str = Query(default="data"),
    sort_dir: str = Query(default="desc", pattern="^(asc|desc)$"),
    user: dict = Depends(get_current_user),
):
    return _get_filtered_history(gerencia, start_date, end_date, search, sort_by, sort_dir)


@router.get("/export/csv")
def export_history_csv(
    gerencia: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    search: str | None = None,
    sort_by: str = Query(default="data"),
    sort_dir: str = Query(default="desc", pattern="^(asc|desc)$"),
    user: dict = Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    items = _get_filtered_history(gerencia, start_date, end_date, search, sort_by, sort_dir)

    def _clean(val: any) -> str:
        s = str(val) if val is not None else ""
        return s.strip()

    STATUS_COLORS = {
        "aprovado":   ("C6EFCE", "276221"),
        "reprovado":  ("FFCCCC", "9C0006"),
        "pendente":   ("FFEB9C", "9C6500"),
        "em análise": ("DDEEFF", "1F4E79"),
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico de Registros"

    HEADERS = [
        "Instalação", "Sistema", "Equipamento", "Gerência", "Data",
        "Situação Identificada", "Status", "Responsável",
        "Data de Criação", "Última Alteração",
    ]

    # ── Cabeçalho ──────────────────────────────────────────────
    header_fill   = PatternFill("solid", fgColor="0B6E4F")
    header_font   = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin          = Side(style="thin", color="CCCCCC")
    cell_border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = cell_border

    ws.row_dimensions[1].height = 32

    # ── Dados ───────────────────────────────────────────────────
    data_font  = Font(name="Calibri", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    left_align = Alignment(vertical="top", horizontal="left")
    center_align = Alignment(vertical="top", horizontal="center")

    fill_even = PatternFill("solid", fgColor="F4F7FB")
    fill_odd  = PatternFill("solid", fgColor="FFFFFF")

    for row_idx, item in enumerate(items, start=2):
        status_raw = _clean(item.get("status") or "Em análise")
        row_data = [
            _clean(item.get("instalacao")),
            _clean(item.get("sistema")),
            _clean(item.get("equipamento")),
            _clean(item.get("gerencia")),
            _clean(item.get("data")),
            _clean(item.get("situacao_identificada")),
            status_raw,
            _clean(item.get("usuario_criacao")),
            _clean(item.get("data_criacao")),
            _clean(item.get("data_alteracao")),
        ]

        is_even = (row_idx % 2 == 0)
        row_fill = fill_even if is_even else fill_odd

        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font   = data_font
            cell.border = cell_border

            # Status cell: cor por valor
            if col == 7:
                sc = STATUS_COLORS.get(value.lower(), ("EEF2F7", "333333"))
                cell.fill = PatternFill("solid", fgColor=sc[0])
                cell.font = Font(name="Calibri", size=10, bold=True, color=sc[1])
                cell.alignment = center_align
            elif col == 5:   # Data
                cell.alignment = center_align
                cell.fill = row_fill
            elif col == 6:   # Situação Identificada – quebra de texto
                cell.alignment = wrap_align
                cell.fill = row_fill
            else:
                cell.alignment = left_align
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = 20

    # ── Larguras das colunas ─────────────────────────────────────
    col_widths = [22, 18, 20, 18, 12, 52, 16, 24, 22, 22]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Filtro automático + painel fixo ─────────────────────────
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # ── Rodapé: gerado em / total de registros ──────────────────
    footer_row = len(items) + 3
    ws.cell(row=footer_row, column=1, value=f"Total de registros: {len(items)}")
    ws.cell(row=footer_row, column=1).font = Font(italic=True, color="888888", size=9)
    ws.cell(row=footer_row + 1, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=footer_row + 1, column=1).font = Font(italic=True, color="888888", size=9)

    # ── Serializar e retornar ────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    filename = quote(f"historico_{timestamp}.xlsx")
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/export/pdf")
def export_history_pdf(
    gerencia: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    search: str | None = None,
    sort_by: str = Query(default="data"),
    sort_dir: str = Query(default="desc", pattern="^(asc|desc)$"),
    user: dict = Depends(get_current_user),
):
    items = _get_filtered_history(gerencia, start_date, end_date, search, sort_by, sort_dir)
    generated_at = datetime.now(timezone.utc).isoformat()
    pdf_content = PDFService().generate_history_pdf(items, user.get("email", ""), generated_at)
    filename = quote(f"historico_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.pdf")
    return StreamingResponse(
        iter([pdf_content]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )
